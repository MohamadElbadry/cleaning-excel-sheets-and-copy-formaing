import os
from openpyxl import load_workbook

# Directory containing the Excel files
folder_path = r'E:\DH\My Work\نشرة اسعار اهم السلعة الغذائيه\Totals\2022\Totals\3 Unpivoted Tables'

# Iterate over each file in the folder
for filename in os.listdir(folder_path):
    if filename.endswith(".xlsx"):  # Check if file is Excel file
        file_path = os.path.join(folder_path, filename)
        
        # Load the workbook
        wb = load_workbook(filename=file_path)
        
        # Iterate over each sheet in the workbook
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # Delete row at index 0 (first row)
            sheet.delete_rows(1, 1)
        
        # Save the modified workbook with the original filename
        wb.save(os.path.join(folder_path, filename))
